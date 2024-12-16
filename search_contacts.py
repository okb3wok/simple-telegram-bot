import openpyxl
import re
from base64 import b64encode
from base64 import b64decode


def search_in_excel(file_path, search_text):
    wb = openpyxl.load_workbook(file_path)
    out = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.rows:
            for cell in row:
                if re.search(r'{}'.format(search_text.lower()), str(cell.value).lower()):
                    out.append({'sheet': sheet_name, 'row': cell.row})
                    break
    return out


def encode_to_base64(text):
    text_bytes = text.encode()
    base64_bytes = b64encode(text_bytes)
    return base64_bytes.decode()


def decode_from_base64(base64):
    base64_bytes = base64.encode()
    message_bytes = b64decode(base64_bytes)
    return message_bytes.decode()


def get_index_by_sheetname(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    return wb.sheetnames.index(sheet_name)


def get_sheetname_by_index(file_path, index):
    wb = openpyxl.load_workbook(file_path)
    return wb.sheetnames[index] 


def parcing_row(file_path, sheet_name, row_number):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    array = list()
    array.append(sheet_name)
    cell_index = 1
    for cell in sheet[row_number]:
        if cell_index < 9:
            if sheet_name == 'Департамент развития' and cell_index == 2:    # left second column
                pass
            elif sheet_name == 'Администрация' and cell_index == 1:  # left first column
                pass
            elif sheet_name == 'Администрация' and cell_index == 5:  # put 'None' they haven`t local phone
                array.append('None')
                array.append(str(cell.value))
            elif sheet_name == 'Служба главного инженера' and cell_index == 4:  # put 'None' they haven`t local phone
                array.append('None')
                array.append(str(cell.value))
            else:
                array.append(str(cell.value))
        cell_index = cell_index + 1
    return array[0:8]


def test_data_array(array):
    result = ''

    if len(array) != 8:
        result = 'Массив'

    if array[7] != 'None' and not re.match(r'^(.*).jpg$', array[7]):
        result = result + ', Фото'

    if array[5] != 'None' and not re.match(r'^(\d{11})$', array[5].replace("-", "").replace(" ", "")):
        result = result + f', Телефон сотовый {array[5]}'

    if array[6] != 'None' and not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', array[6].strip()):
        result = result + f', Емайл {array[6]}'

    return result


def read_row_values(file_path, sheet_name, row_number):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    row_values = ''
    for cell in sheet[row_number]:
        if cell.value:
            matches = re.match(r'^8[\s-]?(\d{3})[\s-]?(\d{3})[\s-]?(\d{2})[\s-]?(\d{2})$', str(cell.value))
            if matches:
                row_values = (row_values + '\n' +
                              f'<a href="+7{matches[1]}{matches[2]}{matches[3]}{matches[4]}">'
                              f'+7{matches[1]}{matches[2]}{matches[3]}{matches[4]}</a>')
            else:
                matchesPhoto = re.match(r'^(.*).jpg$', str(cell.value))
                if matchesPhoto:
                    path = f'{wb.worksheets.index(sheet)}/{row_number}'
                    path_bytes = path.encode()
                    base64_bytes = b64encode(path_bytes)
                    base64_message = base64_bytes.decode()
                    row_values = row_values + '\n' + f'<a href="https://bot-host.site/user/{base64_message}">Анкета</a>'
                else:
                    row_values = row_values + '\n' + str(cell.value)
    return row_values


def read_row_by_sheetnum(file_path, sheet_number, row_number):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[sheet_number]
    array = dict()
    array['phone'] = ''
    array['photo'] = ''
    array['email'] = ''
    array['localphone'] = ''
    array['room'] = ''
    cell_index = 0
    for cell in sheet[row_number]:
        if cell.value:
            matches = re.match(r'^8[\s-]?(\d{3})[\s-]?(\d{3})[\s-]?(\d{2})[\s-]?(\d{2})$', str(cell.value))
            matchesPhoto = re.match(r'^(.*).jpg$', str(cell.value))
            matchesEmail = re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', str(cell.value))
            matchesLocalPhone = re.match(r'^\d{4}$', str(cell.value))


            if matches:
                array['phone'] = f'+7{matches[1]}{matches[2]}{matches[3]}{matches[4]}'
            elif matchesPhoto:
                array['photo'] = f'{matchesPhoto[0]}'
            elif matchesEmail:
                array['email'] = f'{matchesEmail[0]}'
            elif matchesLocalPhone:
                array['localphone'] = f'{matchesLocalPhone[0]}'
            else:
                array[cell_index] = str(cell.value)
            cell_index = cell_index + 1
    return array
